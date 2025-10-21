import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
import pandas as pd
from tabulate import tabulate
import requests
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import schedule
import time
import ssl
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import random  # Импортируем модуль random для генерации случайного времени

# Настройки сервера и User-Agent
SERVER = "https://pepsico.fas.tedo.ru"
USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3"

# Настройки SMTP сервера
SMTP_SERVER = "pepsico.fas.tedo.ru"
SMTP_PORT = 465
SMTP_USERNAME = "monitoring@pepsico.fas.tedo.ru"
SMTP_PASSWORD = "wIr6ZiaQHS5KneqfWbME7Bh0"
EMAIL_FROM = "monitoring@pepsico.fas.tedo.ru"
EMAIL_TO = ["adam.latyrov@tedo.ru"]
EMAIL_SUBJECT = "Еженедельный отчет по производительности Metabase"

# Настройки логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# Функция для отправки письма
def send_email(subject, body, attachment_path=None, retries=3):
    for attempt in range(retries):
        logging.info(f"Попытка отправки email {attempt + 1}/{retries} с темой '{subject}'")
        msg = MIMEMultipart()
        msg['From'] = EMAIL_FROM
        msg['To'] = ", ".join(EMAIL_TO)
        msg['Subject'] = subject

        msg.attach(MIMEText(body, 'plain'))

        if attachment_path:
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f"attachment; filename= {attachment_path.split('/')[-1]}")
                msg.attach(part)

        context = ssl.create_default_context()
        try:
            with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as server:
                server.login(SMTP_USERNAME, SMTP_PASSWORD)
                server.sendmail(EMAIL_FROM, EMAIL_TO, msg.as_string())
                logging.info(f"Email отправлен на {', '.join(EMAIL_TO)} успешно!")
                return
        except Exception as e:
            logging.error(f"Не удалось отправить email: {str(e)}")
            if attempt < retries - 1:
                time.sleep(5)  # задержка перед повторной попыткой
    logging.error(f"Не удалось отправить email после {retries} попыток.")


# Функция для получения токена
def get_token(email, password):
    headers = {
        'Content-Type': 'application/json',
        'User-Agent': USER_AGENT
    }
    logging.info(f"Запрос токена для пользователя: {email}")
    response = requests.post(f"{SERVER}/api/session", json={"username": email, "password": password}, headers=headers)
    response.raise_for_status()
    token = response.json().get('id')
    logging.info(f"Получен токен для пользователя: {email}")
    return token


# Список пользователей и токенов
users = [
    ("metabasetester", "KMiHXovz5LyEa5yxsRI6pKL4R"),
]

tokens = {user[0]: get_token(user[0], user[1]) for user in users}

def run_report(email, token, report_name, report_details, code):
    headers = {
        'Content-Type': 'application/json',
        'User-Agent': USER_AGENT,
        'X-Metabase-Session': token
    }

    updated_parameters = []
    for param in report_details["parameters"]:
        if "company" in param.get("target", "")[1] or "comp" in param.get("target", "")[1]:
            updated_parameters.append({
                **param,
                "value": [code]
            })
        else:
            updated_parameters.append(param)

    updated_details = {
        **report_details,
        "parameters": updated_parameters
    }

    start_time = datetime.now()
    logging.info(
        f"Запуск отчета '{report_name}' для пользователя {email} с параметрами: {[param['value'] for param in updated_details['parameters']]}")
    response = requests.post(updated_details["url"], json=updated_details, headers=headers)
    try:
        response.raise_for_status()
        json_data = response.json()
        if 'error' in json_data:
            error_message = json_data['error']
            logging.error(f"Ошибка выполнения отчета '{report_name}' для пользователя {email}: {error_message}")
            return {'Отчет': report_name, 'Код компании': code,
                    'Параметры': [param['value'] for param in updated_details['parameters']],
                    'Время выполнения (сек)': round((datetime.now() - start_time).total_seconds(), 2),
                    'Ошибка': error_message}
    except requests.exceptions.RequestException as e:
        logging.error(f"Ошибка выполнения отчета '{report_name}' для пользователя {email}: {str(e)}")
        return {'Отчет': report_name, 'Код компании': code,
                'Параметры': [param['value'] for param in updated_details['parameters']],
                'Время выполнения (сек)': round((datetime.now() - start_time).total_seconds(), 2),
                'Ошибка': str(e)}

    end_time = datetime.now()
    execution_time = round((end_time - start_time).total_seconds(), 2)
    logging.info(f"Отчет '{report_name}' для пользователя {email} завершен. Время выполнения: {execution_time} сек")
    return {'Отчет': report_name, 'Код компании': code,
            'Параметры': [param['value'] for param in updated_details['parameters']],
            'Время выполнения (сек)': execution_time}


def run_reports_for_user(email, token, query, code):
    report_name, report_details = query
    result = run_report(email, token, report_name, report_details, code)
    return [result]


# Список запросов
queries = [
    # ("Регламентированные отчеты (форма 1, форма 2) v1",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/21/dashcard/53/card/80/query",
    #   "parameters": [
    #       {"type": "string/=", "value": ["RU03"], "id": "a70b1d0", "target": ["variable", ["template-tag", "comp"]]},
    #       {"type": "date/single", "value": "2023-01-31", "id": "94226b61",
    #        "target": ["variable", ["template-tag", "date"]]},
    #       {"type": "string/=", "value": ["RU01"], "id": "8d8dc526", "target": ["variable", ["template-tag", "versn"]]}],
    #   "dashboard_id": 21
    #   }
    #  ),
    # ("Регламентированные отчеты (форма 1, форма 2) v2",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/22/dashcard/54/card/81/query",
    #   "parameters": [
    #       {"type": "string/=", "value": ["RU03"], "id": "3e76b7b", "target": ["variable", ["template-tag", "company"]]},
    #       {"type": "date/single", "value": "2023-03-08", "id": "3d40058b",
    #        "target": ["variable", ["template-tag", "date"]]}, {"type": "string/=", "value": ["2022"], "id": "63eb572e",
    #                                                            "target": ["variable", ["template-tag", "version"]]}],
    #   "dashboard_id": 22}
    #  ),
    # ("Оборотно сальдовая ведомость по счетам",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/17/dashcard/46/card/73/query",
    #   "parameters":[{"type":"date/single","value":"2023-01-01","id":"72860db6","target":["variable",["template-tag","start_date"]]},{"type":"date/single","value":"2023-01-31","id":"4b402cbe","target":["variable",["template-tag","end_date"]]},{"type":"string/=","value":["RU10"],"id":"3c9da159","target":["variable",["template-tag","comp"]]}],"dashboard_id":17
    #   }),
    # ("Оборотно-сальдовая ведомость по контрагентам",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/17/dashcard/47/card/74/query",
    #   "parameters":[{"type":"date/single","value":"2023-01-01","id":"72860db6","target":["variable",["template-tag","start_date"]]},{"type":"date/single","value":"2023-01-31","id":"4b402cbe","target":["variable",["template-tag","end_date"]]},{"type":"string/=","value":["RU10"],"id":"3c9da159","target":["variable",["template-tag","comp"]]},{"type":"string/=","value":["Поставщик"],"id":"f58c1193","target":["variable",["template-tag","ven_type"]]},{"type":"string/=","value":["-"],"id":"bc353b5d","target":["variable",["template-tag","lifnr"]]},{"type":"string/=","value":["-"],"id":"c9d1da59","target":["variable",["template-tag","kunnr"]]},{"type":"string/=","value":["No"],"id":"a81bef2b","target":["variable",["template-tag","flag"]]}],"dashboard_id":17
    #   }
    #  ),
    # ("Отчет по Запасам",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/25/dashcard/59/card/86/query",
    #   "parameters":[{"type":"string/=","value":["RU03"],"id":"8eae91dd","target":["variable",["template-tag","bukrs"]]},{"type":"string/=","value":["2023"],"id":"26120b1d","target":["variable",["template-tag","year"]]},{"type":"string/=","value":["001"],"id":"b71f1e23","target":["variable",["template-tag","poper"]]},{"type":"string/=","value":["-"],"id":"a7cc424d","target":["variable",["template-tag","matnr"]]}],"dashboard_id":25
    #   }
    #  ),
    # ("Регистр учета ВНА",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/23/dashcard/71/card/112/query",
    #   "parameters":[{"type":"string/=","value":["RU03"],"id":"89e2a4dc","target":["variable",["template-tag","comp"]]},{"type":"string/=","value":["2023"],"id":"ee80025a","target":["variable",["template-tag","gjahr"]]},{"type":"string/=","value":["12"],"id":"f34efa12","target":["variable",["template-tag","peraf"]]},{"type":"string/=","value":["-"],"id":"ac0c5519","target":["variable",["template-tag","asset"]]},{"type":"string/=","value":["-"],"id":"3eea3ab9","target":["variable",["template-tag","as_sub"]]},{"type":"string/=","value":["-"],"id":"f73327a3","target":["variable",["template-tag","as_class"]]},{"type":"string/=","value":["-"],"id":"ddb0eb7","target":["variable",["template-tag","plant"]]}],"dashboard_id":23
    #   }
    #  ),
    # ("Регистр учета РБП",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/24/dashcard/57/card/84/query",
    #   "parameters":[{"type":"string/=","value":["RU03"],"id":"f7f28ca7","target":["variable",["template-tag","comp"]]},{"type":"date/single","value":"2023-01-31","id":"2199ae18","target":["variable",["template-tag","end_date"]]},{"type":"string/=","value":["-"],"id":"85851c06","target":["variable",["template-tag","acc_obj_cat"]]},{"type":"string/=","value":["-"],"id":"b7636ccd","target":["variable",["template-tag","acc_object"]]},{"type":"string/=","value":["-"],"id":"d73c5d2e","target":["variable",["template-tag","acc_obj_num"]]}],"dashboard_id":24
    #   }
    #  ),
    ("Реестр бухгалтерских документов",
     {"url": "https://pepsico.fas.tedo.ru/api/dashboard/18/dashcard/48/card/75/query",
      "parameters":[{"type":"date/single","value":"2023-01-01","id":"d1d13059","target":["variable",["template-tag","start_date"]]},{"type":"date/single","value":"2023-01-02","id":"e0fe5040","target":["variable",["template-tag","end_date"]]},{"type":"string/=","value":["RU10"],"id":"b1059afc","target":["variable",["template-tag","comp"]]},{"type":"string/=","value":["-"],"id":"fd95978e","target":["variable",["template-tag","account"]]},{"type":"string/=","value":["-"],"id":"3ed42146","target":["variable",["template-tag","ven_type"]]},{"type":"string/=","value":["-"],"id":"3007acad","target":["variable",["template-tag","lifnr"]]},{"type":"string/=","value":["-"],"id":"86342389","target":["variable",["template-tag","kunnr"]]},{"type":"string/=","value":["-"],"id":"552db31f","target":["variable",["template-tag","blart"]]},{"type":"string/=","value":["-"],"id":"c648212a","target":["variable",["template-tag","belnr"]]}],"dashboard_id":18
      }
     ),
    #
    # ("Карточка счета",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/19/dashcard/50/card/77/query",
    #   "parameters":[{"type":"string/=","value":["1000061"],"id":"fda9b4e4","target":["variable",["template-tag","account"]]},{"type":"string/=","value":["RU03"],"id":"38450c3e","target":["variable",["template-tag","comp"]]},{"type":"date/single","value":"2023-01-01","id":"a65cdd8c","target":["variable",["template-tag","start_date"]]},{"type":"date/single","value":"2023-01-31","id":"c6d91509","target":["variable",["template-tag","end_date"]]}],"dashboard_id":19
    #   }
    #  ),
    #
    # ("План счетов",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/20/dashcard/51/card/78/query",
    #   "parameters":[{"type":"string/=","value":["RU10"],"id":"50113db7","target":["variable",["template-tag","comp"]]},{"type":"string/=","value":["PI00"],"id":"a3333136","target":["variable",["template-tag","chart"]]}],"dashboard_id":20
    #   }
    #  ),
    # ("Отчет по результатам переоценки на отчетную дату",
    #  {"url": "https://pepsico.fas.tedo.ru/api/dashboard/26/dashcard/64/card/91/query",
    #   "parameters":[{"type":"string/=","value":["RU03"],"id":"14d77ab0","target":["variable",["template-tag","comp"]]},{"type":"date/single","value":"2023-01-31","id":"c44a508a","target":["variable",["template-tag","datum"]]},{"type":"string/=","value":["-"],"id":"38171b25","target":["variable",["template-tag","lifnr"]]},{"type":"string/=","value":["-"],"id":"86442b12","target":["variable",["template-tag","kunnr"]]}],"dashboard_id":26
    #   }
    #  ),
]


def compare_last_two_weeks(row, last_two_weeks, max_results_df):
    current_week_time = row[last_two_weeks[-1]]

    if len(last_two_weeks) < 2:
        return "Недостаточно данных для сравнения"

    previous_week_time = row[last_two_weeks[-2]]

    if pd.notna(previous_week_time):
        if current_week_time > previous_week_time:
            return f"Время по сравнению с предыдущей неделей увеличилось на {round(current_week_time - previous_week_time, 2)} сек"
        elif current_week_time < previous_week_time:
            return f"Время по сравнению с предыдущей неделей уменьшилось на {round(previous_week_time - current_week_time, 2)} сек"
        else:
            return "Время не изменилось"
    else:
        return "Нет данных за предыдущую неделю"


def save_request_details(request_details, week_number, existing_excel_path, writer):
    request_details_df = pd.DataFrame(request_details)
    request_details_df.insert(0, 'Неделя', week_number)

    if os.path.exists(existing_excel_path):
        workbook = load_workbook(existing_excel_path)
        if 'Запросы' in workbook.sheetnames:
            old_data_df = pd.read_excel(existing_excel_path, sheet_name='Запросы')
            updated_df = pd.concat([request_details_df, old_data_df], ignore_index=True)
            workbook.remove(workbook['Запросы'])
        else:
            updated_df = request_details_df
        updated_df.to_excel(writer, sheet_name='Запросы', index=False)
        workbook.save(existing_excel_path)
    else:
        request_details_df.to_excel(writer, sheet_name='Запросы', index=False)


def save_report_generation_time(existing_excel_path, writer):
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    generation_times_df = pd.DataFrame([{'Дата и время': current_datetime, 'Сообщение': 'Отчет сформирован'}])

    if os.path.exists(existing_excel_path):
        workbook = load_workbook(existing_excel_path)
        if 'Время выполнения' in workbook.sheetnames:
            workbook.remove(workbook['Время выполнения'])
        generation_times_df.to_excel(writer, sheet_name='Время выполнения', index=False)
        workbook.save(existing_excel_path)
    else:
        generation_times_df.to_excel(writer, sheet_name='Время выполнения', index=False)

def is_valid_excel_file(file_path):
    try:
        pd.read_excel(file_path, engine='openpyxl')
        return True
    except Exception as e:
        logging.error(f"Ошибка проверки файла Excel: {str(e)}")
        return False


def update_excel(all_results, request_details, existing_excel_path='report_results.xlsx'):
    current_week = f"Неделя {datetime.now().isocalendar()[1]} {datetime.now().year}"
    results_df = pd.DataFrame(all_results)
    results_df['Время выполнения (сек)'] = results_df['Время выполнения (сек)'].round(2)
    max_results_df = results_df.groupby(['Отчет'])['Время выполнения (сек)'].max().reset_index()

    with pd.ExcelWriter(existing_excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        try:
            details_df = pd.read_excel(existing_excel_path, sheet_name='Отчеты')
            logging.info("Существующий файл загружен успешно.")
        except Exception as e:
            logging.error(f"Ошибка загрузки существующего файла Excel: {str(e)}")
            details_df = pd.DataFrame(columns=['Отчет', 'Комментарий', current_week])

        if current_week not in details_df.columns:
            details_df[current_week] = pd.NA

        for _, row in max_results_df.iterrows():
            report_name = row['Отчет']
            max_execution_time = row['Время выполнения (сек)']
            existing_row = details_df[details_df['Отчет'] == report_name]

            if not existing_row.empty:
                index = existing_row.index[0]
                details_df.at[index, current_week] = max_execution_time
                details_df.at[index, 'Комментарий'] = compare_last_two_weeks(details_df.loc[index],
                                                                             details_df.columns[-13:],
                                                                             max_results_df)
                logging.info(f"Обновлена строка для отчета '{report_name}'.")
            else:
                new_row = {
                    'Отчет': report_name,
                    current_week: max_execution_time,
                    'Комментарий': "Нет данных для сравнения"
                }
                details_df = pd.concat([details_df, pd.DataFrame([new_row])], ignore_index=True)
                logging.info(f"Добавлена новая строка для отчета '{report_name}'.")

        columns = details_df.columns.tolist()
        comment_index = columns.index('Комментарий')
        columns.insert(1, columns.pop(comment_index))
        details_df = details_df[columns]

        week_columns = [col for col in details_df.columns if col.startswith('Неделя')]
        if len(week_columns) > 12:
            drop_columns = week_columns[:-12]
            details_df.drop(columns=drop_columns, inplace=True)

        if 'Номер отчета' not in details_df.columns:
            details_df.insert(0, 'Номер отчета', range(1, len(details_df) + 1))
        else:
            details_df['Номер отчета'] = range(1, len(details_df) + 1)

        try:
            details_df.to_excel(writer, sheet_name='Отчеты', index=False)
            logging.info("Файл Excel успешно обновлен.")
        except Exception as e:
            logging.error(f"Ошибка сохранения файла Excel: {str(e)}")

        save_request_details(request_details, current_week, existing_excel_path, writer)
        save_report_generation_time(existing_excel_path, writer)
def run_all_reports():
    all_results = []
    request_details = []
    errors = []
    current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for query in queries:
        results = []
        with ThreadPoolExecutor(max_workers=len(users)) as executor:
            futures = []
            for user in users:
                for code in ["RU03", "RU04", "RU20"]:
                    logging.info(f"Запуск отчета '{query[0]}' для пользователя {user[0]} с кодом компании {code}")
                    future = executor.submit(run_reports_for_user, user[0], tokens[user[0]], query, code)
                    futures.append(future)
            for future in as_completed(futures):
                try:
                    result = future.result()
                    results.extend(result)
                    for res in result:
                        if 'Ошибка' in res:
                            errors.append(res)
                            logging.error(f"Ошибка в результате: {res}")
                        else:
                            request_details.append(res)
                except Exception as e:
                    errors.append(str(e))
                    logging.error(f"Ошибка выполнения: {str(e)}")
        df = pd.DataFrame(results)
        df.sort_values(by=['Отчет'], inplace=True)
        logging.info(f"Результаты выполнения отчетов:\n{tabulate(df, headers='keys', tablefmt='psql')}")
        all_results.extend(results)

    # Вызов функции для обновления данных в Excel файле
    update_excel(all_results, request_details)

    # Отправка отчета по электронной почте
    if errors:
        error_body = "\n".join([str(error) for error in errors])
        send_email("Ошибки выполнения отчетов", error_body)
    else:
        send_email(EMAIL_SUBJECT, "Отчет выполнен успешно. См. вложение.", "report_results.xlsx")


# Планирование еженедельного запуска
def job():
    run_all_reports()


# Первый запуск сразу
run_all_reports()

# Запуск каждую неделю
schedule.every().monday.at("09:00").do(job)

logging.info("Запланировано еженедельное выполнение отчета каждую неделю в понедельник в 09:00")

while True:
    schedule.run_pending()
    time.sleep(1)